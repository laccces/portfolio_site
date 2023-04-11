import sqlite3
from openpyxl import load_workbook

# Load the Excel workbook
workbook_path = 'album_list.xlsx'
workbook = load_workbook(workbook_path)
sheet = workbook.active

# Create and connect to the SQLite database
db_path = 'albums.db'
conn = sqlite3.connect(db_path)
cursor = conn.cursor()

# Create the albums table if it doesn't exist
cursor.execute('''
    CREATE TABLE IF NOT EXISTS albums (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        artist_name TEXT NOT NULL,
        album_name TEXT NOT NULL,
        year INTEGER NOT NULL
    )
''')

# Read data from the Excel sheet and insert it into the database
for row in sheet.iter_rows(min_row=2, values_only=True):  # Start from row 2 to skip the header
    artist_name, album_name, year = row
    cursor.execute('''
        INSERT INTO albums (artist_name, album_name, year)
        VALUES (?, ?, ?)
    ''', (artist_name, album_name, year))


# Commit the changes and close the database connection
conn.commit()
conn.close()

