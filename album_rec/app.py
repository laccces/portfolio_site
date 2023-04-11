from flask import Flask, render_template, request
import sqlite3
import random
import openpyxl
from openpyxl import load_workbook

app = Flask(__name__)
db_path = 'albums.db'

# Initialize the database with the required schema
def init_db():
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS albums (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            artist_name TEXT NOT NULL,
            album_name TEXT NOT NULL,
            year INTEGER NOT NULL,
            reviewed BOOLEAN DEFAULT 0
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS reviews (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            album_id INTEGER NOT NULL,
            rating INTEGER NOT NULL,
            review_text TEXT,
            FOREIGN KEY (album_id) REFERENCES albums (id)
        )
    ''')

    cursor.execute('SELECT COUNT(*) FROM albums')
    num_albums = cursor.fetchone()[0]

    if num_albums == 0:
        # Import data from the Excel file
        wb = openpyxl.load_workbook('album_list.xlsx')
        ws = wb.active
        for row in ws.iter_rows(min_row=2):  # Assuming the first row contains column headers
            artist_name = row[0].value
            album_name = row[1].value
            year = row[2].value
            cursor.execute('''
                INSERT INTO albums (artist_name, album_name, year)
                VALUES (?, ?, ?)
            ''', (artist_name, album_name, year))

    conn.commit()
    conn.close()

init_db()

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/random-album')
def random_album():
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()

    cursor.execute('SELECT COUNT(*) FROM albums WHERE reviewed = 0')
    num_unreviewed_albums = cursor.fetchone()[0]

    if num_unreviewed_albums == 0:
        return {'error': 'No more albums to review.'}

    random_index = random.randint(1, num_unreviewed_albums)
    cursor.execute('SELECT id, artist_name, album_name, year FROM albums WHERE reviewed = 0 LIMIT 1 OFFSET ?', (random_index - 1,))

    album = cursor.fetchone()
    conn.close()

    return {'id': album[0], 'artist_name': album[1], 'album_name': album[2], 'year': album[3]}

@app.route('/submit-review', methods=['POST'])
def submit_review():
    album_id = request.form.get('album_id')
    rating = request.form.get('rating')
    review_text = request.form.get('review_text')

    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()

    cursor.execute('''
        INSERT INTO reviews (album_id, rating, review_text)
        VALUES (?, ?, ?)
    ''', (album_id, rating, review_text))

    cursor.execute('UPDATE albums SET reviewed = 1 WHERE id = ?', (album_id,))

    conn.commit()
    conn.close()

    return {'success': True}

@app.route('/reviewed-albums')
def reviewed_albums():
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row  # Enables accessing columns by name
    cursor = conn.cursor()

    cursor.execute('''
        SELECT a.artist_name, a.album_name, a.year, r.rating, r.review_text
        FROM albums a
        JOIN reviews r ON a.id = r.album_id
        WHERE a.reviewed = 1
        ORDER BY a.artist_name, a.album_name
    ''')

    albums = cursor.fetchall()
    conn.close()

    return render_template('reviewed_albums.html', albums=albums)


if __name__ == '__main__':
    app.run(debug=True)
